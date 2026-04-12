#!/usr/bin/env python3
"""
Excel to JSON Artifact Converter for Slay the Robot
Converts artifact configurations from Excel to JSON format
"""

import pandas as pd
import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Tuple
from dataclasses import dataclass

# Configuration
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
EXCEL_FILE = PROJECT_ROOT / "external/config/artifacts.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "external/data/artifacts"

# Valid rarity values (matches ArtifactData.ARTIFACT_RARITIES)
VALID_RARITIES = {0: "STARTER", 1: "COMMON", 2: "UNCOMMON", 3: "RARE", 4: "BOSS", 5: "SPECIAL"}

# Valid color IDs
VALID_COLORS = [
    "color_white", "color_red", "color_green", "color_blue",
    "color_purple", "color_orange", "color_yellow"
]


@dataclass
class ValidationError:
    artifact_id: str
    field: str
    message: str
    severity: str = "ERROR"


class ArtifactValidator:
    """Validates artifact configuration data"""

    def __init__(self):
        self.errors: List[ValidationError] = []

    def validate_artifact(self, row: pd.Series) -> List[ValidationError]:
        """Validate a single artifact row"""
        self.errors = []
        artifact_id = str(row.get('object_id', 'UNKNOWN')).strip()

        # Skip empty rows or comment rows
        if not artifact_id or artifact_id.startswith('#') or artifact_id in ['说明', '注释', 'COMMENT', '备注', 'NOTE']:
            return self.errors

        # Required fields
        self._check_required(row, 'object_id', artifact_id)
        self._check_required(row, 'artifact_name', artifact_id)
        self._check_required(row, 'artifact_description', artifact_id)

        # Validate rarity
        self._validate_rarity(row, artifact_id)

        # Validate color
        self._validate_color(row, artifact_id)

        return self.errors

    def _check_required(self, row: pd.Series, field: str, artifact_id: str):
        if field not in row or pd.isna(row[field]) or str(row[field]).strip() == '':
            self.errors.append(ValidationError(
                artifact_id=artifact_id,
                field=field,
                message=f"Required field '{field}' is empty",
                severity="ERROR"
            ))

    def _validate_rarity(self, row: pd.Series, artifact_id: str):
        if 'artifact_rarity' not in row or pd.isna(row['artifact_rarity']):
            return

        rarity = row['artifact_rarity']
        valid_rarities = list(VALID_RARITIES.keys())

        try:
            rarity_val = int(rarity)
            if rarity_val not in valid_rarities:
                self.errors.append(ValidationError(
                    artifact_id=artifact_id,
                    field='artifact_rarity',
                    message=f"Invalid rarity '{rarity}'. Must be one of: {valid_rarities}",
                    severity="ERROR"
                ))
        except (ValueError, TypeError):
            self.errors.append(ValidationError(
                artifact_id=artifact_id,
                field='artifact_rarity',
                message=f"artifact_rarity must be a number, got: {rarity}",
                severity="ERROR"
            ))

    def _validate_color(self, row: pd.Series, artifact_id: str):
        if 'artifact_color_id' not in row or pd.isna(row['artifact_color_id']):
            return

        color = row['artifact_color_id']
        if color not in VALID_COLORS:
            self.errors.append(ValidationError(
                artifact_id=artifact_id,
                field='artifact_color_id',
                message=f"Unknown color '{color}'. Valid colors: {VALID_COLORS}",
                severity="WARNING"
            ))


class ArtifactConverter:
    """Converts Excel data to JSON format"""

    def __init__(self):
        self.validator = ArtifactValidator()
        self.all_errors: List[ValidationError] = []

    def convert(self) -> Tuple[bool, List[ValidationError]]:
        """Main conversion function"""
        print("=" * 60)
        print("Excel to JSON Artifact Converter")
        print("=" * 60)

        # Check Excel file exists
        if not EXCEL_FILE.exists():
            print(f"ERROR: Excel file not found: {EXCEL_FILE}")
            print("Please create the Excel file first.")
            return False, []

        # Create output directory
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Read Excel
        try:
            print(f"\nReading Excel: {EXCEL_FILE}")
            df = pd.read_excel(EXCEL_FILE, sheet_name="Artifacts")
            print(f"Found {len(df)} artifacts")
        except Exception as e:
            print(f"ERROR reading Excel: {e}")
            return False, []

        # Validate and convert each artifact
        success_count = 0
        skipped_rows = 0
        self.all_errors = []

        for idx, row in df.iterrows():
            artifact_id = str(row.get('object_id', f'ROW_{idx}')).strip()

            # Skip empty rows or comment rows
            if not artifact_id or artifact_id.startswith('#') or artifact_id in ['说明', '注释', 'COMMENT', '备注', 'NOTE']:
                skipped_rows += 1
                continue

            print(f"\n[{idx+1}/{len(df)}] Processing: {artifact_id}")

            # Validate
            errors = self.validator.validate_artifact(row)
            self.all_errors.extend(errors)

            if errors:
                for err in errors:
                    prefix = "[WARN]" if err.severity == "WARNING" else "[ERROR]"
                    print(f"  {prefix}: [{err.field}] {err.message}")

            # Skip artifacts with critical errors
            critical_errors = [e for e in errors if e.severity == "ERROR"]
            if critical_errors:
                print(f"  [SKIP] Skipped due to {len(critical_errors)} error(s)")
                continue

            # Convert to JSON
            try:
                artifact_json = self._build_artifact_json(row)
                output_path = OUTPUT_DIR / f"{artifact_id}.json"

                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(artifact_json, f, indent=4, ensure_ascii=False)

                print(f"  [OK] Generated: {output_path}")
                success_count += 1

            except Exception as e:
                print(f"  [FAIL] Conversion failed: {e}")
                import traceback
                traceback.print_exc()
                self.all_errors.append(ValidationError(
                    artifact_id=artifact_id,
                    field="CONVERSION",
                    message=str(e),
                    severity="ERROR"
                ))

        # Summary
        actual_artifacts = len(df) - skipped_rows
        print("\n" + "=" * 60)
        print(f"转换完成!")
        print(f"  成功: {success_count}/{actual_artifacts} 遗物")
        if skipped_rows > 0:
            print(f"  跳过: {skipped_rows} 说明行")
        print(f"  错误: {len([e for e in self.all_errors if e.severity == 'ERROR'])}")
        print(f"  警告: {len([e for e in self.all_errors if e.severity == 'WARNING'])}")
        print("=" * 60)

        return success_count > 0, self.all_errors

    def _build_artifact_json(self, row: pd.Series) -> Dict[str, Any]:
        """Build JSON structure from Excel row"""
        artifact_id = str(row['object_id']).strip()

        # Build actions
        add_actions = self._parse_actions(row, 'artifact_add_action')
        remove_actions = self._parse_actions(row, 'artifact_remove_action')
        turn_start_actions = self._parse_actions(row, 'artifact_turn_start_action')
        turn_end_actions = self._parse_actions(row, 'artifact_turn_end_action')
        first_turn_actions = self._parse_actions(row, 'artifact_first_turn_action')
        end_of_combat_actions = self._parse_actions(row, 'artifact_end_of_combat_action')
        max_counter_actions = self._parse_actions(row, 'artifact_max_counter_action')
        right_click_actions = self._parse_actions(row, 'artifact_right_click_action')

        # Build validators
        right_click_validators = self._parse_validators(row, 'artifact_right_click_validator')

        # Build the final JSON
        json_data = {
            "patch_data": {},
            "properties": {
                "object_id": artifact_id,
                "object_uid": "",
                "artifact_name": self._get_str(row, 'artifact_name'),
                "artifact_description": self._get_str(row, 'artifact_description'),
                "artifact_rarity": self._get_int(row, 'artifact_rarity', 1),
                "artifact_color_id": self._get_str(row, 'artifact_color_id', 'color_white'),
                "artifact_texture_path": self._get_str(row, 'artifact_texture_path', ''),
                "artifact_script_path": self._get_str(row, 'artifact_script_path', 'res://scripts/artifacts/BaseArtifact.gd'),
                "artifact_counter": self._get_int(row, 'artifact_counter', 0),
                "artifact_counter_max": self._get_int(row, 'artifact_counter_max', 1),
                "artifact_counter_reset_on_combat_end": self._get_int(row, 'artifact_counter_reset_on_combat_end', -1),
                "artifact_counter_reset_on_turn_start": self._get_int(row, 'artifact_counter_reset_on_turn_start', -1),
                "artifact_counter_wraparound": self._get_bool(row, 'artifact_counter_wraparound', True),
                "artifact_appears_in_artifact_packs": self._get_bool(row, 'artifact_appears_in_artifact_packs', True),
                "artifact_add_actions": add_actions,
                "artifact_remove_actions": remove_actions,
                "artifact_turn_start_actions": turn_start_actions,
                "artifact_turn_end_actions": turn_end_actions,
                "artifact_first_turn_actions": first_turn_actions,
                "artifact_end_of_combat_actions": end_of_combat_actions,
                "artifact_max_counter_actions": max_counter_actions,
                "artifact_right_click_actions": right_click_actions,
                "artifact_right_click_validators": right_click_validators,
            }
        }

        return json_data

    def _get_str(self, row: pd.Series, field: str, default: str = '') -> str:
        """Get string value with NaN handling"""
        if field not in row or pd.isna(row[field]):
            return default
        val = str(row[field])
        if val.lower() == 'nan':
            return default
        return val

    def _get_int(self, row: pd.Series, field: str, default: int = 0) -> int:
        """Get int value with NaN handling"""
        if field not in row or pd.isna(row[field]):
            return default
        try:
            return int(row[field])
        except (ValueError, TypeError):
            return default

    def _get_bool(self, row: pd.Series, field: str, default: bool = False) -> bool:
        """Get bool value with NaN handling"""
        if field not in row or pd.isna(row[field]):
            return default
        val = row[field]
        if isinstance(val, bool):
            return val
        if isinstance(val, str):
            return val.lower() in ('true', 'yes', '1')
        return bool(val)

    # Action path mappings
    ACTION_PATHS = {
        # meta_actions
        'ActionAttackGenerator': 'meta_actions',
        'ActionCardPlay': 'meta_actions',
        'ActionCardPlayEnd': 'meta_actions',
        'ActionDrawGenerator': 'meta_actions',
        'ActionEmitCustomSignal': 'meta_actions',
        'ActionValidator': 'meta_actions',
        'ActionVariableCardsetModifier': 'meta_actions',
        'ActionVariableCombatStatsModifier': 'meta_actions',
        'ActionVariableCostModifier': 'meta_actions',
        # artifact_actions
        'ActionIncreaseArtifactCharge': 'artifact_actions',
        # cardset_actions
        'ActionAddCardsToDeck': 'cardset_actions',
        'ActionAddCardsToDraw': 'cardset_actions',
        'ActionAddCardsToHand': 'cardset_actions',
        'ActionAttachCardsOntoEnemy': 'cardset_actions',
        'ActionBanishCards': 'cardset_actions',
        'ActionAttachCardToSlot': 'cardset_actions',
        'ActionChangeCardEnergies': 'cardset_actions',
        'ActionChangeCardProperties': 'cardset_actions',
        'ActionDiscardCards': 'cardset_actions',
        'ActionExhaustCards': 'cardset_actions',
        'ActionImproveCardValues': 'cardset_actions',
        'ActionMoveCardsToLimbo': 'cardset_actions',
        'ActionPlayCards': 'cardset_actions',
        'ActionRandomizeCardEnergies': 'cardset_actions',
        'ActionRemoveCardsFromDeck': 'cardset_actions',
        'ActionRetainCards': 'cardset_actions',
        'ActionTransformCards': 'cardset_actions',
        'ActionUpgradeCards': 'cardset_actions',
        # custom_ui_actions
        'ActionCustomUI': 'custom_ui_actions',
        # debug_actions
        'ActionDebugLog': 'debug_actions',
        # enemy_actions
        'ActionCycleEnemyIntent': 'enemy_actions',
        # pick_card_actions
        'ActionBasePickCards': 'pick_card_actions',
        'ActionCreateCards': 'pick_card_actions',
        'ActionPickCards': 'pick_card_actions',
        'ActionPickDuplicateCards': 'pick_card_actions',
        'ActionPickUpgradeCards': 'pick_card_actions',
        # player_actions
        'ActionAddArtifact': 'player_actions',
        'ActionAddConsumable': 'player_actions',
        'ActionAddMoney': 'player_actions',
        'ActionSwapBossArtifact': 'player_actions',
        'ActionUpdateCardDrafts': 'player_actions',
        'ActionUpdateRestActions': 'player_actions',
        'ActionUseConsumable': 'player_actions',
        # status_actions
        'ActionApplyStatus': 'status_actions',
        'ActionCorrosion': 'status_actions',
        'ActionDecayStatus': 'status_actions',
        # world_interaction_actions
        'ActionOpenChest': 'world_interaction_actions',
        'ActionStartCombat': 'world_interaction_actions',
        'ActionVisitLocation': 'world_interaction_actions',
        # generated_actions
        'ActionAttack': 'generated_actions',
        'ActionDraw': 'generated_actions',
        # rewards
        'ActionClearRewards': 'rewards',
        'ActionGrantRewards': 'rewards',
        # shop_actions
        'ActionShopPopulateItems': 'shop_actions',
        'ActionShopPurchaseItems': 'shop_actions',
        # world_generation_actions
        'ActionGenerateAct': 'world_generation_actions',
    }

    def _get_action_path(self, action_type: str) -> str:
        """Get the correct path for an action type"""
        if not action_type.startswith(("Action", "Artifact")):
            return f"res://scripts/actions/{action_type}.gd"

        subdir = self.ACTION_PATHS.get(action_type, '')
        if subdir:
            return f"res://scripts/actions/{subdir}/{action_type}.gd"
        else:
            return f"res://scripts/actions/{action_type}.gd"

    def _parse_actions(self, row: pd.Series, action_prefix: str) -> List[Dict]:
        """Parse action configuration from row"""
        actions = []

        # Get action type
        action_type_col = f"{action_prefix}_type"
        if action_type_col not in row or pd.isna(row[action_type_col]):
            return actions

        action_type = str(row[action_type_col]).strip()
        if not action_type or action_type.lower() == 'nan':
            return actions

        # Build action parameters
        params = {}

        # Common parameters
        time_delay_col = f"{action_prefix}_time_delay"
        if time_delay_col in row and pd.notna(row[time_delay_col]):
            try:
                params['time_delay'] = float(row[time_delay_col])
            except (ValueError, TypeError):
                pass

        target_override_col = f"{action_prefix}_target_override"
        if target_override_col in row and pd.notna(row[target_override_col]):
            params['target_override'] = str(row[target_override_col])

        # Action-specific parameters
        block_col = f"{action_prefix}_block"
        if block_col in row and pd.notna(row[block_col]):
            try:
                params['block'] = float(row[block_col])
            except (ValueError, TypeError):
                pass

        money_col = f"{action_prefix}_money_amount"
        if money_col in row and pd.notna(row[money_col]):
            try:
                params['money_amount'] = float(row[money_col])
            except (ValueError, TypeError):
                pass

        # Build final action with correct path
        action_path = self._get_action_path(action_type)
        if action_type.startswith(("Action", "Artifact")):
            actions.append({action_path: params})

        return actions

    def _parse_validators(self, row: pd.Series, validator_prefix: str) -> List[Dict]:
        """Parse validator configuration from row"""
        validators = []

        # Get validator type
        validator_type_col = f"{validator_prefix}_type"
        if validator_type_col not in row or pd.isna(row[validator_type_col]):
            # Default validator for right_click
            if 'right_click' in validator_prefix:
                return [{"res://scripts/validators/ValidatorPlayerTurn.gd": {}}]
            return validators

        validator_type = str(row[validator_type_col]).strip()
        if not validator_type or validator_type.lower() == 'nan':
            if 'right_click' in validator_prefix:
                return [{"res://scripts/validators/ValidatorPlayerTurn.gd": {}}]
            return validators

        # Build validator parameters
        params = {}

        # Build final validator
        validator_path = f"res://scripts/validators/{validator_type}.gd"
        validators.append({validator_path: params})

        return validators


def create_sample_excel():
    """Create a sample Excel file with proper structure"""
    print("\n创建示例遗物Excel文件...")
    print(f"输出位置: {EXCEL_FILE}")

    # Define columns
    columns = [
        # Basic Info
        'object_id', 'artifact_name', 'artifact_description', 'artifact_rarity', 'artifact_color_id',

        # Visual
        'artifact_texture_path',

        # Script
        'artifact_script_path',

        # Counter
        'artifact_counter', 'artifact_counter_max',
        'artifact_counter_reset_on_combat_end', 'artifact_counter_reset_on_turn_start',
        'artifact_counter_wraparound', 'artifact_appears_in_artifact_packs',

        # Actions - Add
        'artifact_add_action_type', 'artifact_add_action_money_amount',

        # Actions - Turn Start
        'artifact_turn_start_action_type', 'artifact_turn_start_action_block',

        # Actions - Max Counter
        'artifact_max_counter_action_type', 'artifact_max_counter_action_block',

        # Actions - Right Click
        'artifact_right_click_action_type',
    ]

    # Sample data
    sample_data = [
        {
            'object_id': '# 说明行（此行会被自动跳过）',
            'artifact_name': '在下方添加你的遗物数据',
            'artifact_description': 'artifact_rarity: 0=初始,1=普通,2=稀有,3=罕见,4=Boss,5=特殊 | artifact_color_id: color_white/red/green/blue/purple/orange/yellow',
        },
        {
            'object_id': 'artifact_add_money',
            'artifact_name': 'Add Money',
            'artifact_description': 'Adds money when obtained',
            'artifact_rarity': 1,  # COMMON
            'artifact_color_id': 'color_white',
            'artifact_texture_path': 'external/sprites/artifacts/artifact_white.png',
            'artifact_script_path': 'res://scripts/artifacts/BaseArtifact.gd',
            'artifact_counter': 0,
            'artifact_counter_max': 1,
            'artifact_counter_reset_on_combat_end': -1,
            'artifact_counter_reset_on_turn_start': -1,
            'artifact_counter_wraparound': True,
            'artifact_appears_in_artifact_packs': True,
            'artifact_add_action_type': 'ActionAddMoney',
            'artifact_add_action_money_amount': 200,
        },
        {
            'object_id': 'artifact_block_on_attacks',
            'artifact_name': 'Block on Attacks',
            'artifact_description': 'Grants 5 block every 3 attacks',
            'artifact_rarity': 1,  # COMMON
            'artifact_color_id': 'color_red',
            'artifact_texture_path': 'external/sprites/artifacts/artifact_red.png',
            'artifact_script_path': 'res://scripts/artifacts/ArtifactBlockOnAttacks.gd',
            'artifact_counter': 0,
            'artifact_counter_max': 3,
            'artifact_counter_reset_on_combat_end': 0,
            'artifact_counter_reset_on_turn_start': 0,
            'artifact_counter_wraparound': True,
            'artifact_appears_in_artifact_packs': True,
            'artifact_max_counter_action_type': 'ActionBlock',
            'artifact_max_counter_action_block': 5,
        },
    ]

    # Create DataFrame
    df = pd.DataFrame(sample_data, columns=columns)

    # Write Excel with formatting
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Artifacts', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Artifacts']

        # Add formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Header formatting
        header_fill = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"[OK] 示例遗物Excel已创建: {EXCEL_FILE}")
    print(f"     包含 {len(sample_data)-1} 个示例遗物 (第1行为说明行)")


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Excel to JSON Artifact Converter')
    parser.add_argument('--create-sample', action='store_true',
                       help='Create a sample Excel file')
    parser.add_argument('--validate-only', action='store_true',
                       help='Only validate, do not generate JSON')

    args = parser.parse_args()

    if args.create_sample:
        create_sample_excel()
        return

    # Run conversion
    converter = ArtifactConverter()
    success, errors = converter.convert()

    if not success and not errors:
        print("\nNo Excel file found. Create a sample?")
        print(f"  python {sys.argv[0]} --create-sample")
        sys.exit(1)

    critical_errors = [e for e in errors if e.severity == "ERROR"]
    sys.exit(0 if len(critical_errors) == 0 else 1)


if __name__ == "__main__":
    main()
