#!/usr/bin/env python3
"""
Excel to JSON Event Converter for Slay the Robot
Converts event configurations from Excel to JSON format
"""

import pandas as pd
import json
import sys
import re
from pathlib import Path
from typing import Dict, List, Any, Tuple
from dataclasses import dataclass

# Configuration
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
EXCEL_FILE = PROJECT_ROOT / "external/config/events.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "external/data/events"

# Valid strategies for validator failed
VALID_FAILED_STRATEGIES = {
    0: "RETRY_ONCE",
    1: "PICK_ANYWAY",
    2: "REMOVE_FROM_POOL"
}


@dataclass
class ValidationError:
    event_id: str
    field: str
    message: str
    severity: str = "ERROR"


class EventValidator:
    """Validates event configuration data"""

    def __init__(self):
        self.errors: List[ValidationError] = []

    def validate_event(self, row: pd.Series) -> List[ValidationError]:
        """Validate a single event row"""
        self.errors = []
        event_id = str(row.get('object_id', 'UNKNOWN')).strip()

        # Skip empty rows or comment rows
        if not event_id or event_id.startswith('#') or event_id in ['说明', '注释', 'COMMENT', '备注', 'NOTE']:
            return self.errors

        # Required fields
        self._check_required(row, 'object_id', event_id)

        # Validate failed strategy
        self._validate_failed_strategy(row, event_id)

        # Validate enemy placement positions format
        self._validate_positions(row, event_id)

        # Validate weighted enemies format
        self._validate_weighted_enemies(row, event_id)

        return self.errors

    def _check_required(self, row: pd.Series, field: str, event_id: str):
        if field not in row or pd.isna(row[field]) or str(row[field]).strip() == '':
            self.errors.append(ValidationError(
                event_id=event_id,
                field=field,
                message=f"Required field '{field}' is empty",
                severity="ERROR"
            ))

    def _validate_failed_strategy(self, row: pd.Series, event_id: str):
        if 'location_event_pool_validator_failed_strategy' not in row or pd.isna(row['location_event_pool_validator_failed_strategy']):
            return

        strategy = row['location_event_pool_validator_failed_strategy']
        valid_strategies = list(VALID_FAILED_STRATEGIES.keys())

        try:
            strategy_val = int(strategy)
            if strategy_val not in valid_strategies:
                self.errors.append(ValidationError(
                    event_id=event_id,
                    field='location_event_pool_validator_failed_strategy',
                    message=f"Invalid strategy '{strategy}'. Must be one of: {valid_strategies} (RETRY_ONCE=0, PICK_ANYWAY=1, REMOVE_FROM_POOL=2)",
                    severity="WARNING"
                ))
        except (ValueError, TypeError):
            self.errors.append(ValidationError(
                event_id=event_id,
                field='location_event_pool_validator_failed_strategy',
                message=f"Strategy must be a number, got: {strategy}",
                severity="ERROR"
            ))

    def _validate_positions(self, row: pd.Series, event_id: str):
        """Validate enemy placement positions format"""
        if 'event_enemy_placement_positions' not in row or pd.isna(row['event_enemy_placement_positions']):
            return

        positions_str = str(row['event_enemy_placement_positions']).strip()
        if not positions_str or positions_str.lower() == 'nan' or positions_str == 'auto':
            return

        # Check format: should be like "0,-40|0,40" or "0,-40;0,40"
        # Each position should have x,y coordinates
        try:
            # Try to parse as pipe or semicolon separated positions
            positions = re.split(r'[|;]', positions_str)
            for pos in positions:
                coords = pos.strip().split(',')
                if len(coords) != 2:
                    self.errors.append(ValidationError(
                        event_id=event_id,
                        field='event_enemy_placement_positions',
                        message=f"Invalid position format '{pos}'. Expected 'x,y' (e.g., '0,-40')",
                        severity="WARNING"
                    ))
                    return
                float(coords[0])
                float(coords[1])
        except (ValueError, IndexError):
            self.errors.append(ValidationError(
                event_id=event_id,
                field='event_enemy_placement_positions',
                message=f"Invalid positions format. Expected format: 'x1,y1|x2,y2' (e.g., '0,-40|0,40')",
                severity="WARNING"
            ))

    def _validate_weighted_enemies(self, row: pd.Series, event_id: str):
        """Validate weighted enemy object IDs format"""
        if 'event_weighted_enemy_object_ids' not in row or pd.isna(row['event_weighted_enemy_object_ids']):
            return

        enemies_str = str(row['event_weighted_enemy_object_ids']).strip()
        if not enemies_str or enemies_str.lower() == 'nan':
            return

        # Check format: should be like "enemy_1:1,enemy_2:1|enemy_1:1"
        # Groups separated by |, enemies within group separated by ,
        try:
            groups = re.split(r'[|]', enemies_str)
            for group in groups:
                enemies = group.strip().split(',')
                for enemy in enemies:
                    if ':' in enemy:
                        parts = enemy.strip().split(':')
                        if len(parts) != 2:
                            raise ValueError(f"Invalid enemy format: {enemy}")
                        float(parts[1])  # Weight should be a number
        except (ValueError, IndexError):
            self.errors.append(ValidationError(
                event_id=event_id,
                field='event_weighted_enemy_object_ids',
                message=f"Invalid enemies format. Expected format: 'enemy_id:weight,enemy_id:weight|enemy_id:weight'",
                severity="WARNING"
            ))


class EventConverter:
    """Converts Excel data to JSON format"""

    def __init__(self):
        self.validator = EventValidator()
        self.all_errors: List[ValidationError] = []

    def convert(self) -> Tuple[bool, List[ValidationError]]:
        """Main conversion function"""
        print("=" * 60)
        print("Excel to JSON Event Converter")
        print("=" * 60)

        # Check Excel file exists
        if not EXCEL_FILE.exists():
            print(f"ERROR: Excel file not found: {EXCEL_FILE}")
            print("Please create the Excel file first.")
            return False, []

        # Create output directory
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Read Excel
        try:
            print(f"\nReading Excel: {EXCEL_FILE}")
            df = pd.read_excel(EXCEL_FILE, sheet_name="Events")
            print(f"Found {len(df)} events")
        except Exception as e:
            print(f"ERROR reading Excel: {e}")
            return False, []

        # Validate and convert each event
        success_count = 0
        skipped_rows = 0
        self.all_errors = []

        for idx, row in df.iterrows():
            event_id = str(row.get('object_id', f'ROW_{idx}')).strip()

            # Skip empty rows or comment rows
            if not event_id or event_id.startswith('#') or event_id in ['说明', '注释', 'COMMENT', '备注', 'NOTE']:
                skipped_rows += 1
                continue

            print(f"\n[{idx+1}/{len(df)}] Processing: {event_id}")

            # Validate
            errors = self.validator.validate_event(row)
            self.all_errors.extend(errors)

            if errors:
                for err in errors:
                    prefix = "[WARN]" if err.severity == "WARNING" else "[ERROR]"
                    print(f"  {prefix}: [{err.field}] {err.message}")

            # Skip events with critical errors
            critical_errors = [e for e in errors if e.severity == "ERROR"]
            if critical_errors:
                print(f"  [SKIP] Skipped due to {len(critical_errors)} error(s)")
                continue

            # Convert to JSON
            try:
                event_json = self._build_event_json(row)
                output_path = OUTPUT_DIR / f"{event_id}.json"

                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(event_json, f, indent=4, ensure_ascii=False)

                print(f"  [OK] Generated: {output_path}")
                success_count += 1

            except Exception as e:
                print(f"  [FAIL] Conversion failed: {e}")
                import traceback
                traceback.print_exc()
                self.all_errors.append(ValidationError(
                    event_id=event_id,
                    field="CONVERSION",
                    message=str(e),
                    severity="ERROR"
                ))

        # Summary
        actual_events = len(df) - skipped_rows
        print("\n" + "=" * 60)
        print(f"转换完成!")
        print(f"  成功: {success_count}/{actual_events} 事件")
        if skipped_rows > 0:
            print(f"  跳过: {skipped_rows} 说明行")
        print(f"  错误: {len([e for e in self.all_errors if e.severity == 'ERROR'])}")
        print(f"  警告: {len([e for e in self.all_errors if e.severity == 'WARNING'])}")
        print("=" * 60)

        return success_count > 0, self.all_errors

    def _build_event_json(self, row: pd.Series) -> Dict[str, Any]:
        """Build JSON structure from Excel row"""
        event_id = str(row['object_id']).strip()

        # Parse enemy placement positions
        positions = self._parse_positions(row)

        # Parse weighted enemy object IDs
        weighted_enemies = self._parse_weighted_enemies(row)

        # Build the final JSON
        json_data = {
            "patch_data": {},
            "properties": {
                "object_id": event_id,
                "event_background_texture_path": self._get_str(row, 'event_background_texture_path', ''),
                "event_dialogue_object_id": self._get_str(row, 'event_dialogue_object_id', ''),
                "event_enemy_placement_is_automatic": self._get_bool(row, 'event_enemy_placement_is_automatic', True),
                "event_enemy_placement_positions": positions,
                "event_initial_combat_actions": [],  # Complex structure, skip for now
                "event_pool_validator_data": [],  # Complex structure, skip for now
                "event_weighted_enemy_object_ids": weighted_enemies,
                "location_event_pool_validator_failed_strategy": self._get_int(
                    row, 'location_event_pool_validator_failed_strategy', 1
                ),
            }
        }

        return json_data

    def _get_str(self, row: pd.Series, field: str, default: str = '') -> str:
        """Get string value with NaN handling"""
        if field not in row or pd.isna(row[field]):
            return default
        val = str(row[field])
        if val.lower() == 'nan':
            return default
        return val

    def _get_int(self, row: pd.Series, field: str, default: int = 0) -> int:
        """Get int value with NaN handling"""
        if field not in row or pd.isna(row[field]):
            return default
        try:
            return int(row[field])
        except (ValueError, TypeError):
            return default

    def _get_bool(self, row: pd.Series, field: str, default: bool = False) -> bool:
        """Get bool value with NaN handling"""
        if field not in row or pd.isna(row[field]):
            return default
        val = row[field]
        if isinstance(val, bool):
            return val
        if isinstance(val, str):
            return val.lower() in ('true', 'yes', '1', 'automatic')
        return bool(val)

    def _parse_positions(self, row: pd.Series) -> List[List[float]]:
        """Parse enemy placement positions"""
        if 'event_enemy_placement_positions' not in row or pd.isna(row['event_enemy_placement_positions']):
            return []

        positions_str = str(row['event_enemy_placement_positions']).strip()
        if not positions_str or positions_str.lower() == 'nan' or positions_str == 'auto':
            return []

        positions = []
        try:
            # Format: "x1,y1|x2,y2" or "x1,y1;x2,y2"
            pos_groups = re.split(r'[|;]', positions_str)
            for pos in pos_groups:
                coords = pos.strip().split(',')
                if len(coords) == 2:
                    x = float(coords[0].strip())
                    y = float(coords[1].strip())
                    positions.append([x, y])
        except (ValueError, IndexError):
            pass

        return positions

    def _parse_weighted_enemies(self, row: pd.Series) -> List[Dict[str, float]]:
        """Parse weighted enemy object IDs"""
        if 'event_weighted_enemy_object_ids' not in row or pd.isna(row['event_weighted_enemy_object_ids']):
            return []

        enemies_str = str(row['event_weighted_enemy_object_ids']).strip()
        if not enemies_str or enemies_str.lower() == 'nan':
            return []

        weighted_enemies = []
        try:
            # Format: "enemy_1:1,enemy_2:1|enemy_1:1" (groups separated by |)
            groups = re.split(r'[|]', enemies_str)
            for group in groups:
                enemy_dict = {}
                enemies = group.strip().split(',')
                for enemy in enemies:
                    enemy = enemy.strip()
                    if ':' in enemy:
                        parts = enemy.split(':')
                        enemy_id = parts[0].strip()
                        weight = float(parts[1].strip())
                        enemy_dict[enemy_id] = weight
                    else:
                        # No weight specified, default to 1.0
                        enemy_dict[enemy] = 1.0
                if enemy_dict:
                    weighted_enemies.append(enemy_dict)
        except (ValueError, IndexError):
            pass

        return weighted_enemies


def create_sample_excel():
    """Create a sample Excel file with proper structure"""
    print("\n创建示例事件Excel文件...")
    print(f"输出位置: {EXCEL_FILE}")

    # Define columns
    columns = [
        # Basic Info
        'object_id', 'event_dialogue_object_id', 'event_background_texture_path',

        # Placement
        'event_enemy_placement_is_automatic', 'event_enemy_placement_positions',

        # Enemies
        'event_weighted_enemy_object_ids',

        # Strategy
        'location_event_pool_validator_failed_strategy',
    ]

    # Sample data
    sample_data = [
        {
            'object_id': '# 说明行（此行会被自动跳过）',
            'event_dialogue_object_id': '在下方添加你的事件数据',
            'event_background_texture_path': 'event_enemy_placement_positions格式: x,y|x,y (多位置用|分隔) | event_weighted_enemy_object_ids格式: enemy_id:weight,enemy_id:weight|enemy_id:weight (多组用|分隔)',
        },
        {
            'object_id': 'event_act_1_boss_1',
            'event_dialogue_object_id': '',
            'event_background_texture_path': '',
            'event_enemy_placement_is_automatic': False,
            'event_enemy_placement_positions': '0,0|180,0|360,0',
            'event_weighted_enemy_object_ids': 'enemy_act_1_boss_1:1',
            'location_event_pool_validator_failed_strategy': 1,
        },
        {
            'object_id': 'event_act_1_easy_combat_1',
            'event_dialogue_object_id': '',
            'event_background_texture_path': '',
            'event_enemy_placement_is_automatic': True,
            'event_enemy_placement_positions': '0,-40|0,40',
            'event_weighted_enemy_object_ids': 'enemy_1:1,enemy_2:1,enemy_3:1|enemy_1:1,enemy_2:1,enemy_3:1|enemy_1:1,enemy_2:1,enemy_3:1',
            'location_event_pool_validator_failed_strategy': 1,
        },
        {
            'object_id': 'event_pick_something',
            'event_dialogue_object_id': 'dialogue_pick_something',
            'event_background_texture_path': '',
            'event_enemy_placement_is_automatic': True,
            'event_enemy_placement_positions': '0,-40|0,40',
            'event_weighted_enemy_object_ids': 'enemy_1:1,enemy_2:1|enemy_2:1',
            'location_event_pool_validator_failed_strategy': 1,
        },
    ]

    # Create DataFrame
    df = pd.DataFrame(sample_data, columns=columns)

    # Write Excel with formatting
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Events', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Events']

        # Add formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Header formatting
        header_fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"[OK] 示例事件Excel已创建: {EXCEL_FILE}")
    print(f"     包含 {len(sample_data)-1} 个示例事件 (第1行为说明行)")


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Excel to JSON Event Converter')
    parser.add_argument('--create-sample', action='store_true',
                       help='Create a sample Excel file')
    parser.add_argument('--validate-only', action='store_true',
                       help='Only validate, do not generate JSON')

    args = parser.parse_args()

    if args.create_sample:
        create_sample_excel()
        return

    # Run conversion
    converter = EventConverter()
    success, errors = converter.convert()

    if not success and not errors:
        print("\nNo Excel file found. Create a sample?")
        print(f"  python {sys.argv[0]} --create-sample")
        sys.exit(1)

    critical_errors = [e for e in errors if e.severity == "ERROR"]
    sys.exit(0 if len(critical_errors) == 0 else 1)


if __name__ == "__main__":
    main()
