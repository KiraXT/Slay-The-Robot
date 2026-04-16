#!/usr/bin/env python3
"""
JSON to Excel Converter for Slay the Robot
Converts existing JSON configurations back to Excel format
"""

import pandas as pd
import json
import os
import re
from pathlib import Path
from typing import Dict, List, Any, Optional

# Configuration
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

CARDS_DIR = PROJECT_ROOT / "external/data/cards"
ARTIFACTS_DIR = PROJECT_ROOT / "external/data/artifacts"
EVENTS_DIR = PROJECT_ROOT / "external/data/events"

CARDS_EXCEL = PROJECT_ROOT / "external/config/cards.xlsx"
ARTIFACTS_EXCEL = PROJECT_ROOT / "external/config/artifacts.xlsx"
EVENTS_EXCEL = PROJECT_ROOT / "external/config/events.xlsx"


def load_json_files(directory: Path) -> List[Dict]:
    """Load all JSON files from directory"""
    data = []
    if not directory.exists():
        return data
    for file_path in sorted(directory.glob("*.json")):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = json.load(f)
                if isinstance(content, dict) and "properties" in content:
                    data.append(content["properties"])
                elif isinstance(content, dict):
                    data.append(content)
        except Exception as e:
            print(f"[WARN] Failed to load {file_path}: {e}")
    return data


def extract_action_name(action_path: str) -> str:
    """Extract action class name from path like res://scripts/actions/.../ActionName.gd"""
    if not action_path:
        return ""
    return Path(action_path).stem


def parse_simple_action(actions: List[Dict]) -> Dict[str, Any]:
    """Parse first simple action into Excel columns"""
    result = {
        "action_type": "",
        "action_time_delay": "",
        "action_target_override": "",
        "action_on_lethal": "",
        "validator_type": "",
    }
    if not actions or not isinstance(actions, list):
        return result

    first_action = actions[0]
    if not isinstance(first_action, dict):
        return result

    for path, params in first_action.items():
        result["action_type"] = extract_action_name(path)
        if isinstance(params, dict):
            if "time_delay" in params:
                result["action_time_delay"] = params["time_delay"]
            if "target_override" in params:
                to = params["target_override"]
                if isinstance(to, int):
                    target_map = {
                        0: "BaseAction.TARGET_OVERRIDES.SELECTED_TARGETS",
                        1: "BaseAction.TARGET_OVERRIDES.PARENT",
                        2: "BaseAction.TARGET_OVERRIDES.PLAYER",
                        3: "BaseAction.TARGET_OVERRIDES.ALL_COMBATANTS",
                        4: "BaseAction.TARGET_OVERRIDES.ALL_ENEMIES",
                        5: "BaseAction.TARGET_OVERRIDES.LEFTMOST_ENEMY",
                        6: "BaseAction.TARGET_OVERRIDES.ENEMY_ID",
                        7: "BaseAction.TARGET_OVERRIDES.RANDOM_ENEMY",
                    }
                    result["action_target_override"] = target_map.get(to, to)
                else:
                    result["action_target_override"] = to
            if "actions_on_lethal" in params and params["actions_on_lethal"]:
                lethal = params["actions_on_lethal"]
                if isinstance(lethal, list) and lethal:
                    lethal_path = list(lethal[0].keys())[0]
                    result["action_on_lethal"] = extract_action_name(lethal_path)
            if "validator_data" in params and params["validator_data"]:
                validators = params["validator_data"]
                if isinstance(validators, list) and validators:
                    validator_path = list(validators[0].keys())[0]
                    result["validator_type"] = extract_action_name(validator_path)
        break

    return result


def parse_artifact_actions(actions: List[Dict], prefix: str) -> Dict[str, Any]:
    """Parse artifact actions into prefixed Excel columns"""
    result = {
        f"{prefix}_type": "",
        f"{prefix}_block": "",
        f"{prefix}_money_amount": "",
    }
    if not actions or not isinstance(actions, list):
        return result

    first_action = actions[0]
    if not isinstance(first_action, dict):
        return result

    for path, params in first_action.items():
        result[f"{prefix}_type"] = extract_action_name(path)
        if isinstance(params, dict):
            if "block" in params:
                result[f"{prefix}_block"] = params["block"]
            if "money_amount" in params:
                result[f"{prefix}_money_amount"] = params["money_amount"]
        break

    return result


def format_positions(positions: List[List[float]]) -> str:
    """Format positions as 'x,y|x,y'"""
    if not positions:
        return ""
    return "|".join(f"{p[0]},{p[1]}" for p in positions)


def format_weighted_enemies(enemies: List[Dict[str, float]]) -> str:
    """Format weighted enemies as 'enemy_id:weight,enemy_id:weight|enemy_id:weight'"""
    if not enemies:
        return ""
    groups = []
    for group in enemies:
        items = [f"{k}:{v}" for k, v in group.items()]
        groups.append(",".join(items))
    return "|".join(groups)


def convert_cards() -> None:
    """Convert card JSONs to Excel"""
    print("=" * 60)
    print("Converting Cards JSON -> Excel")
    print("=" * 60)

    cards = load_json_files(CARDS_DIR)
    print(f"Loaded {len(cards)} cards")

    columns = [
        'object_id', 'card_name', 'card_description', 'card_type', 'card_rarity',
        'card_color_id', 'card_energy_cost', 'card_energy_cost_is_variable',
        'card_energy_cost_variable_upper_bound', 'card_requires_target',
        'card_exhausts', 'card_is_ethereal', 'card_is_retained',
        'card_appears_in_card_packs', 'card_texture_path', 'card_keyword_object_ids',
        'damage', 'number_of_attacks', 'block', 'draw_count',
        'status_charge_amount', 'status_secondary_charge_amount',
        'money_amount', 'heal_amount', 'damage_random',
        'status_effect_object_id', 'status_force_apply_new_effect',
        'random_consumable', 'fill_all_slots', 'ignored_interceptor_ids',
        'multiplier_offset',
        'action_type', 'action_time_delay', 'action_target_override',
        'action_on_lethal', 'validator_type',
        'card_upgrade_amount_max', 'first_upgrade_energy_cost',
        'upgrade_damage', 'upgrade_block', 'upgrade_number_of_attacks',
        'upgrade_draw_count', 'upgrade_status_charge_amount',
        'upgrade_status_secondary_charge_amount', 'upgrade_damage_random',
        'upgrade_multiplier_offset', 'card_first_shuffle_priority',
    ]

    # Instruction row
    rows = [{
        'object_id': '# 说明行（此行会被自动跳过）',
        'card_name': '在下方添加你的卡牌数据',
        'card_description': 'card_type: 0=攻击,1=技能,2=能力,3=状态,4=诅咒 | card_rarity: 0=基础,1=普通,2=稀有,3=罕见,4=生成',
    }]

    for card in cards:
        row = {}
        for col in columns:
            row[col] = ""

        row['object_id'] = card.get('object_id', '')
        row['card_name'] = card.get('card_name', '')
        row['card_description'] = card.get('card_description', '')
        row['card_type'] = card.get('card_type', '')
        row['card_rarity'] = card.get('card_rarity', '')
        row['card_color_id'] = card.get('card_color_id', '')
        row['card_energy_cost'] = card.get('card_energy_cost', '')
        row['card_energy_cost_is_variable'] = card.get('card_energy_cost_is_variable', False)
        row['card_energy_cost_variable_upper_bound'] = card.get('card_energy_cost_variable_upper_bound', -1)
        row['card_requires_target'] = card.get('card_requires_target', True)
        row['card_exhausts'] = card.get('card_exhausts', False)
        row['card_is_ethereal'] = card.get('card_is_ethereal', False)
        row['card_is_retained'] = card.get('card_is_retained', False)
        row['card_appears_in_card_packs'] = card.get('card_appears_in_card_packs', True)
        row['card_texture_path'] = card.get('card_texture_path', '')
        row['card_first_shuffle_priority'] = card.get('card_first_shuffle_priority', 0)
        row['card_upgrade_amount_max'] = card.get('card_upgrade_amount_max', 1)

        keywords = card.get('card_keyword_object_ids', [])
        row['card_keyword_object_ids'] = ",".join(keywords) if keywords else ""

        card_values = card.get('card_values', {})
        for key, val in card_values.items():
            if key in columns:
                row[key] = val

        actions = parse_simple_action(card.get('card_play_actions', []))
        row.update(actions)

        upgrades = card.get('card_upgrade_value_improvements', {})
        for key, val in upgrades.items():
            excel_key = f"upgrade_{key}"
            if excel_key in columns:
                row[excel_key] = val

        first_upgrade = card.get('card_first_upgrade_property_changes', {})
        if 'card_energy_cost' in first_upgrade:
            row['first_upgrade_energy_cost'] = first_upgrade['card_energy_cost']

        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    CARDS_EXCEL.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(CARDS_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cards', index=False)
        _format_worksheet(writer.sheets['Cards'], '366092')

    print(f"[OK] Cards Excel saved: {CARDS_EXCEL} ({len(rows)-1} cards)")


def convert_artifacts() -> None:
    """Convert artifact JSONs to Excel"""
    print("=" * 60)
    print("Converting Artifacts JSON -> Excel")
    print("=" * 60)

    artifacts = load_json_files(ARTIFACTS_DIR)
    print(f"Loaded {len(artifacts)} artifacts")

    columns = [
        'object_id', 'artifact_name', 'artifact_description', 'artifact_rarity', 'artifact_color_id',
        'artifact_texture_path', 'artifact_script_path',
        'artifact_counter', 'artifact_counter_max',
        'artifact_counter_reset_on_combat_end', 'artifact_counter_reset_on_turn_start',
        'artifact_counter_wraparound', 'artifact_appears_in_artifact_packs',
        'artifact_add_action_type', 'artifact_add_action_money_amount',
        'artifact_remove_action_type', 'artifact_remove_action_money_amount',
        'artifact_turn_start_action_type', 'artifact_turn_start_action_block',
        'artifact_turn_end_action_type', 'artifact_turn_end_action_block',
        'artifact_first_turn_action_type', 'artifact_first_turn_action_block',
        'artifact_end_of_combat_action_type', 'artifact_end_of_combat_action_block',
        'artifact_max_counter_action_type', 'artifact_max_counter_action_block',
        'artifact_max_counter_action_money_amount',
        'artifact_right_click_action_type', 'artifact_right_click_action_block',
        'artifact_right_click_validator_type',
    ]

    rows = [{
        'object_id': '# 说明行（此行会被自动跳过）',
        'artifact_name': '在下方添加你的遗物数据',
        'artifact_description': 'artifact_rarity: 0=初始,1=普通,2=稀有,3=罕见,4=Boss,5=特殊 | artifact_color_id: color_white/red/green/blue/purple/orange/yellow',
    }]

    for artifact in artifacts:
        row = {}
        for col in columns:
            row[col] = ""

        row['object_id'] = artifact.get('object_id', '')
        row['artifact_name'] = artifact.get('artifact_name', '')
        row['artifact_description'] = artifact.get('artifact_description', '')
        row['artifact_rarity'] = artifact.get('artifact_rarity', 1)
        row['artifact_color_id'] = artifact.get('artifact_color_id', 'color_white')
        row['artifact_texture_path'] = artifact.get('artifact_texture_path', '')
        row['artifact_script_path'] = artifact.get('artifact_script_path', 'res://scripts/artifacts/BaseArtifact.gd')
        row['artifact_counter'] = artifact.get('artifact_counter', 0)
        row['artifact_counter_max'] = artifact.get('artifact_counter_max', 1)
        row['artifact_counter_reset_on_combat_end'] = artifact.get('artifact_counter_reset_on_combat_end', -1)
        row['artifact_counter_reset_on_turn_start'] = artifact.get('artifact_counter_reset_on_turn_start', -1)
        row['artifact_counter_wraparound'] = artifact.get('artifact_counter_wraparound', True)
        row['artifact_appears_in_artifact_packs'] = artifact.get('artifact_appears_in_artifact_packs', True)

        row.update(parse_artifact_actions(artifact.get('artifact_add_actions', []), 'artifact_add_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_remove_actions', []), 'artifact_remove_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_turn_start_actions', []), 'artifact_turn_start_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_turn_end_actions', []), 'artifact_turn_end_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_first_turn_actions', []), 'artifact_first_turn_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_end_of_combat_actions', []), 'artifact_end_of_combat_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_max_counter_actions', []), 'artifact_max_counter_action'))
        row.update(parse_artifact_actions(artifact.get('artifact_right_click_actions', []), 'artifact_right_click_action'))

        validators = artifact.get('artifact_right_click_validators', [])
        if validators and isinstance(validators, list):
            v_path = list(validators[0].keys())[0]
            row['artifact_right_click_validator_type'] = extract_action_name(v_path)

        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    ARTIFACTS_EXCEL.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(ARTIFACTS_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Artifacts', index=False)
        _format_worksheet(writer.sheets['Artifacts'], '8B4513')

    print(f"[OK] Artifacts Excel saved: {ARTIFACTS_EXCEL} ({len(rows)-1} artifacts)")


def convert_events() -> None:
    """Convert event JSONs to Excel"""
    print("=" * 60)
    print("Converting Events JSON -> Excel")
    print("=" * 60)

    events = load_json_files(EVENTS_DIR)
    print(f"Loaded {len(events)} events")

    columns = [
        'object_id', 'event_dialogue_object_id', 'event_background_texture_path',
        'event_enemy_placement_is_automatic', 'event_enemy_placement_positions',
        'event_weighted_enemy_object_ids',
        'location_event_pool_validator_failed_strategy',
    ]

    rows = [{
        'object_id': '# 说明行（此行会被自动跳过）',
        'event_dialogue_object_id': '在下方添加你的事件数据',
        'event_background_texture_path': 'event_enemy_placement_positions格式: x,y|x,y (多位置用|分隔) | event_weighted_enemy_object_ids格式: enemy_id:weight,enemy_id:weight|enemy_id:weight (多组用|分隔)',
    }]

    for event in events:
        row = {}
        for col in columns:
            row[col] = ""

        row['object_id'] = event.get('object_id', '')
        row['event_dialogue_object_id'] = event.get('event_dialogue_object_id', '')
        row['event_background_texture_path'] = event.get('event_background_texture_path', '')
        row['event_enemy_placement_is_automatic'] = event.get('event_enemy_placement_is_automatic', True)
        row['event_enemy_placement_positions'] = format_positions(event.get('event_enemy_placement_positions', []))
        row['event_weighted_enemy_object_ids'] = format_weighted_enemies(event.get('event_weighted_enemy_object_ids', []))
        row['location_event_pool_validator_failed_strategy'] = event.get('location_event_pool_validator_failed_strategy', 1)

        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    EVENTS_EXCEL.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(EVENTS_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Events', index=False)
        _format_worksheet(writer.sheets['Events'], '2E8B57')

    print(f"[OK] Events Excel saved: {EVENTS_EXCEL} ({len(rows)-1} events)")


def _format_worksheet(worksheet, header_color: str):
    """Apply header formatting and auto-width"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    convert_cards()
    print()
    convert_artifacts()
    print()
    convert_events()
    print()
    print("=" * 60)
    print("All Excel files have been updated from JSON!")
    print("=" * 60)


if __name__ == "__main__":
    main()
