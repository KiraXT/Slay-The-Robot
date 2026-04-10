#!/usr/bin/env python3
"""
Extract card data from Global.gd add_test_cards() function
and create a complete Excel configuration file
"""

import re
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any

GLOBAL_GD_PATH = Path("autoload/Global.gd")
OUTPUT_EXCEL = Path("external/config/cards.xlsx")

def parse_card_data(file_path: Path) -> List[Dict[str, Any]]:
    """Parse card definitions from Global.gd"""

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the add_test_cards function
    func_match = re.search(
        r'func add_test_cards\(\).*?(?=func \w+\(|\Z)',
        content,
        re.DOTALL
    )

    if not func_match:
        print("ERROR: Could not find add_test_cards() function")
        return []

    func_content = func_match.group(0)

    # Find all card definitions
    # Pattern: var card_name: CardData = CardData.new("object_id")
    card_pattern = r'var\s+(\w+)\s*:\s*CardData\s*=\s*CardData\.new\("([^"]+)"\)(.*?)(?=var\s+\w+\s*:\s*CardData|\Z)'
    cards = re.findall(card_pattern, func_content, re.DOTALL)

    parsed_cards = []

    for var_name, object_id, card_body in cards:
        card_data = {
            'object_id': object_id,
            'var_name': var_name,
        }

        # Extract all property assignments
        # Pattern: card_name.property = value
        prop_pattern = rf'{re.escape(var_name)}\.(\w+)\s*=\s*(.+?)(?:\n|$)'
        properties = re.findall(prop_pattern, card_body)

        for prop_name, prop_value in properties:
            prop_value = prop_value.strip()

            # Clean up GDScript syntax
            if prop_value.startswith('CardData.CARD_TYPES.'):
                card_data['card_type'] = parse_card_type(prop_value)
            elif prop_value.startswith('CardData.CARD_RARITIES.'):
                card_data['card_rarity'] = parse_rarity(prop_value)
            elif prop_name == 'card_values':
                card_data['card_values'] = parse_dict(prop_value)
            elif prop_name == 'card_upgrade_value_improvements':
                card_data['upgrade_improvements'] = parse_dict(prop_value)
            elif prop_name == 'card_first_upgrade_property_changes':
                card_data['first_upgrade_changes'] = parse_dict(prop_value)
            elif prop_name == 'card_play_actions':
                card_data['actions'] = parse_actions(prop_value)
            elif prop_name == 'card_end_of_turn_actions':
                card_data['end_of_turn_actions'] = parse_actions(prop_value)
            elif prop_name == 'card_retain_actions':
                card_data['retain_actions'] = parse_actions(prop_value)
            elif prop_name == 'card_initial_combat_actions':
                card_data['initial_combat_actions'] = parse_actions(prop_value)
            elif prop_name == 'card_keyword_object_ids':
                card_data['keywords'] = parse_array(prop_value)
            elif prop_name in ['card_is_retained', 'card_exhausts', 'card_is_ethereal',
                              'card_requires_target', 'card_energy_cost_is_variable',
                              'card_appears_in_card_packs']:
                card_data[prop_name] = parse_bool(prop_value)
            elif prop_name in ['card_energy_cost', 'card_upgrade_amount_max',
                              'card_first_shuffle_priority', 'card_energy_cost_variable_upper_bound']:
                try:
                    card_data[prop_name] = int(prop_value)
                except:
                    card_data[prop_name] = prop_value
            else:
                # String values
                if prop_value.startswith('"') and prop_value.endswith('"'):
                    card_data[prop_name] = prop_value[1:-1]
                else:
                    card_data[prop_name] = prop_value

        # Flatten card_values into individual columns
        if 'card_values' in card_data and isinstance(card_data['card_values'], dict):
            for key, value in card_data['card_values'].items():
                card_data[key] = value

        # Flatten upgrade improvements
        if 'upgrade_improvements' in card_data and isinstance(card_data['upgrade_improvements'], dict):
            for key, value in card_data['upgrade_improvements'].items():
                card_data[f'upgrade_{key}'] = value

        # Flatten first upgrade changes
        if 'first_upgrade_changes' in card_data and isinstance(card_data['first_upgrade_changes'], dict):
            for key, value in card_data['first_upgrade_changes'].items():
                if key == 'card_energy_cost':
                    card_data['first_upgrade_energy_cost'] = value

        # Extract action info
        if 'actions' in card_data and card_data['actions']:
            action_info = card_data['actions'][0] if isinstance(card_data['actions'], list) else card_data['actions']
            if isinstance(action_info, dict):
                action_script = list(action_info.keys())[0] if action_info else ''
                action_params = list(action_info.values())[0] if action_info else {}

                # Extract action type from script path
                action_type = action_script.split('/')[-1].replace('.gd', '') if action_script else ''
                card_data['action_type'] = action_type

                # Extract parameters
                if isinstance(action_params, dict):
                    card_data['action_time_delay'] = action_params.get('time_delay', 0.0)
                    if 'target_override' in action_params:
                        card_data['action_target_override'] = action_params['target_override']
                    if 'actions_on_lethal' in action_params and action_params['actions_on_lethal']:
                        card_data['action_on_lethal'] = 'ActionAddMoney'  # Simplified

        parsed_cards.append(card_data)

    return parsed_cards


def parse_card_type(value: str) -> int:
    """Convert GDScript card type to integer"""
    type_map = {
        'CardData.CARD_TYPES.ATTACK': 0,
        'CardData.CARD_TYPES.SKILL': 1,
        'CardData.CARD_TYPES.POWER': 2,
        'CardData.CARD_TYPES.STATUS': 3,
        'CardData.CARD_TYPES.CURSE': 4,
    }
    return type_map.get(value, 0)


def parse_rarity(value: str) -> int:
    """Convert GDScript rarity to integer"""
    rarity_map = {
        'CardData.CARD_RARITIES.BASIC': 0,
        'CardData.CARD_RARITIES.COMMON': 1,
        'CardData.CARD_RARITIES.UNCOMMON': 2,
        'CardData.CARD_RARITIES.RARE': 3,
        'CardData.CARD_RARITIES.GENERATED': 4,
    }
    return rarity_map.get(value, 1)


def parse_dict(value: str) -> Dict:
    """Parse a GDScript dictionary"""
    result = {}
    # Remove outer braces
    value = value.strip()
    if value.startswith('{'):
        value = value[1:]
    if value.endswith('}'):
        value = value[:-1]

    # Simple key-value parsing
    # Pattern: "key": value or key: value
    pairs = re.findall(r'"([^"]+)"\s*:\s*([^,\n]+)', value)
    for key, val in pairs:
        val = val.strip()
        # Try to convert to int
        try:
            result[key] = int(val)
        except ValueError:
            # Check for boolean
            if val.lower() == 'true':
                result[key] = True
            elif val.lower() == 'false':
                result[key] = False
            elif val.startswith('['):
                # Array value
                result[key] = parse_array(val)
            elif val.startswith('"') and val.endswith('"'):
                result[key] = val[1:-1]
            else:
                result[key] = val

    return result


def parse_array(value: str) -> List:
    """Parse a GDScript array"""
    value = value.strip()
    if value.startswith('['):
        value = value[1:]
    if value.endswith(']'):
        value = value[:-1]

    if not value or value.strip() == '':
        return []

    items = []
    for item in re.split(r',\s*(?![^\[]*\])', value):
        item = item.strip()
        if item.startswith('"') and item.endswith('"'):
            items.append(item[1:-1])
        elif item:
            items.append(item)

    return items


def parse_bool(value: str) -> bool:
    """Parse GDScript boolean"""
    return value.lower() == 'true'


def parse_actions(value: str) -> List[Dict]:
    """Parse GDScript action arrays"""
    actions = []

    # Find action script references
    # Pattern: Scripts.ACTION_XXX: { params }
    action_pattern = r'Scripts\.(ACTION_\w+)\s*:\s*(\{[^}]*\})'
    matches = re.findall(action_pattern, value)

    for action_name, params_str in matches:
        action_dict = {}
        params = parse_dict(params_str)

        # Convert to script path
        script_name = action_name.replace('ACTION_', '').replace('_', '').title()
        script_name = script_name.replace('Generator', 'Generator')  # Keep Generator suffix

        # Special handling for action names
        script_map = {
            'ActionAttackGenerator': 'ActionAttackGenerator',
            'ActionBlock': 'ActionBlock',
            'ActionDrawGenerator': 'ActionDrawGenerator',
            'ActionApplyStatus': 'ActionApplyStatus',
            'ActionReshuffle': 'ActionReshuffle',
            'ActionAddConsumable': 'ActionAddConsumable',
            'ActionAddMoney': 'ActionAddMoney',
            'ActionAddHealth': 'ActionAddHealth',
            'ActionValidator': 'ActionValidator',
            'ActionAttachCardsOntoEnemy': 'ActionAttachCardsOntoEnemy',
            'ActionImproveCardValues': 'ActionImproveCardValues',
            'ActionVariableCostModifier': 'ActionVariableCostModifier',
        }

        script_name = script_map.get(action_name.replace('ACTION_', 'Action'), script_name)
        script_path = f"res://scripts/actions/{script_name}.gd"

        action_dict[script_path] = params
        actions.append(action_dict)

    return actions


def flatten_to_excel(cards: List[Dict]) -> pd.DataFrame:
    """Convert parsed cards to a flat DataFrame for Excel"""

    # Define all columns
    columns = [
        # Basic Info
        'object_id', 'card_name', 'card_description', 'card_type', 'card_rarity',
        'card_color_id', 'card_energy_cost', 'card_energy_cost_is_variable',
        'card_energy_cost_variable_upper_bound', 'card_requires_target',

        # Flags
        'card_exhausts', 'card_is_ethereal', 'card_is_retained',
        'card_appears_in_card_packs',

        # Visual
        'card_texture_path', 'card_keyword_object_ids',

        # Card Values
        'damage', 'number_of_attacks', 'block', 'draw_count',
        'status_charge_amount', 'status_secondary_charge_amount',
        'money_amount', 'heal_amount', 'damage_random',
        'status_effect_object_id', 'status_force_apply_new_effect',
        'random_consumable', 'fill_all_slots', 'ignored_interceptor_ids',
        'multiplier_offset',

        # Actions
        'action_type', 'action_time_delay', 'action_target_override',
        'action_on_lethal', 'validator_type',

        # Upgrades
        'card_upgrade_amount_max', 'first_upgrade_energy_cost',
        'upgrade_damage', 'upgrade_block', 'upgrade_number_of_attacks',
        'upgrade_draw_count', 'upgrade_status_charge_amount',
        'upgrade_status_secondary_charge_amount', 'upgrade_damage_random',
        'upgrade_multiplier_offset',

        # Shuffle
        'card_first_shuffle_priority',
    ]

    # Normalize each card to have all columns
    rows = []
    for card in cards:
        row = {}
        for col in columns:
            value = card.get(col, '')
            # Convert lists to comma-separated strings
            if isinstance(value, list):
                value = ', '.join(str(v) for v in value)
            row[col] = value
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    return df


def create_formatted_excel(df: pd.DataFrame, output_path: Path):
    """Create an Excel file with formatting and data validation"""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cards', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Cards']

        # Import openpyxl components
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Freeze header row
        worksheet.freeze_panes = 'A2'

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            # Set width with min/max bounds
            adjusted_width = max(12, min(max_length + 2, 40))
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add data validation for enums
        # Card Type validation
        type_validation = DataValidation(
            type="list",
            formula1='"0-ATTACK,1-SKILL,2-POWER,3-STATUS,4-CURSE"',
            allow_blank=True
        )
        type_validation.error = 'Please select a valid card type'
        type_validation.errorTitle = 'Invalid Card Type'
        worksheet.add_data_validation(type_validation)

        # Find card_type column and apply validation
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'card_type':
                col_letter = get_column_letter(idx)
                type_validation.add(f'{col_letter}2:{col_letter}{len(df)+100}')
                break

        # Rarity validation
        rarity_validation = DataValidation(
            type="list",
            formula1='"0-BASIC,1-COMMON,2-UNCOMMON,3-RARE,4-GENERATED"',
            allow_blank=True
        )
        worksheet.add_data_validation(rarity_validation)

        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'card_rarity':
                col_letter = get_column_letter(idx)
                rarity_validation.add(f'{col_letter}2:{col_letter}{len(df)+100}')
                break

        # Color validation
        color_validation = DataValidation(
            type="list",
            formula1='"color_white,color_red,color_green,color_blue,color_purple,color_orange,color_yellow"',
            allow_blank=True
        )
        worksheet.add_data_validation(color_validation)

        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'card_color_id':
                col_letter = get_column_letter(idx)
                color_validation.add(f'{col_letter}2:{col_letter}{len(df)+100}')
                break

        # Boolean validation for flag columns
        bool_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True
        )
        worksheet.add_data_validation(bool_validation)

        bool_columns = ['card_exhausts', 'card_is_ethereal', 'card_is_retained',
                       'card_requires_target', 'card_energy_cost_is_variable',
                       'card_appears_in_card_packs']

        for idx, col_name in enumerate(df.columns, 1):
            if col_name in bool_columns:
                col_letter = get_column_letter(idx)
                bool_validation.add(f'{col_letter}2:{col_letter}{len(df)+100}')

        # Add alternating row colors for readability
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row_idx in range(2, len(df) + 2, 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = light_fill

        # Set row height for header
        worksheet.row_dimensions[1].height = 30

    print(f"✅ Excel file created: {output_path}")
    print(f"   Total cards: {len(df)}")


def main():
    print("=" * 60)
    print("Extracting Cards from Global.gd")
    print("=" * 60)

    if not GLOBAL_GD_PATH.exists():
        print(f"ERROR: Global.gd not found at {GLOBAL_GD_PATH}")
        print("Please run this script from the project root directory")
        return

    print(f"\nParsing: {GLOBAL_GD_PATH}")
    cards = parse_card_data(GLOBAL_GD_PATH)

    if not cards:
        print("No cards found!")
        return

    print(f"Found {len(cards)} cards")

    # Flatten to DataFrame
    df = flatten_to_excel(cards)

    # Create formatted Excel
    create_formatted_excel(df, OUTPUT_EXCEL)

    print("\n" + "=" * 60)
    print("Next steps:")
    print("  1. Open: external/config/cards.xlsx")
    print("  2. Edit card data as needed")
    print("  3. Run: python external/tools/excel_to_json.py")
    print("  4. Or double-click: external/tools/convert.bat")
    print("=" * 60)


if __name__ == "__main__":
    main()
