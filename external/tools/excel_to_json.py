#!/usr/bin/env python3
"""
Excel to JSON Card Converter for Slay the Robot
Converts card configurations from Excel to JSON format
Includes validation rules to prevent configuration errors
"""

import pandas as pd
import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from enum import IntEnum

# Configuration
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
EXCEL_FILE = PROJECT_ROOT / "external/config/cards.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "external/data/cards"

# Card Type Enum (matches CardData.CARD_TYPES)
class CardType(IntEnum):
    ATTACK = 0
    SKILL = 1
    POWER = 2
    STATUS = 3
    CURSE = 4

# Card Rarity Enum (matches CardData.CARD_RARITIES)
class CardRarity(IntEnum):
    BASIC = 0
    COMMON = 1
    UNCOMMON = 2
    RARE = 3
    GENERATED = 4

# Valid color IDs
VALID_COLORS = [
    "color_white", "color_red", "color_green", "color_blue",
    "color_purple", "color_orange", "color_yellow"
]

# Valid keyword IDs (should be populated from your game data)
VALID_KEYWORDS = [
    "keyword_block", "keyword_corrosion", "keyword_bomb",
    "keyword_vulnerable", "keyword_weaken", "keyword_preserve_block"
]

# Valid action script paths
VALID_ACTIONS = [
    "ActionAttackGenerator", "ActionBlock", "ActionDrawGenerator",
    "ActionApplyStatus", "ActionReshuffle", "ActionAddConsumable",
    "ActionAddMoney", "ActionAddHealth", "ActionValidator",
    "ActionAttachCardsOntoEnemy", "ActionImproveCardValues",
    "ActionVariableCostModifier", "ActionDirectDamage", "ActionEndTurn"
]

# Valid validators
VALID_VALIDATORS = [
    "ValidatorCardTypeInHand", "ValidatorEnemyAttacking",
    "ValidatorCardPlayEnemyAttacking", "ValidatorPlayerTurn"
]


@dataclass
class ValidationError:
    card_id: str
    field: str
    message: str
    severity: str = "ERROR"  # ERROR or WARNING


class CardValidator:
    """Validates card configuration data"""

    def __init__(self):
        self.errors: List[ValidationError] = []

    def validate_card(self, row: pd.Series) -> List[ValidationError]:
        """Validate a single card row"""
        self.errors = []
        card_id = str(row.get('object_id', 'UNKNOWN'))

        # Required fields
        self._check_required(row, 'object_id', card_id)
        self._check_required(row, 'card_name', card_id)
        self._check_required(row, 'card_type', card_id)
        self._check_required(row, 'card_description', card_id)

        # Type validation
        self._validate_card_type(row, card_id)
        self._validate_rarity(row, card_id)
        self._validate_color(row, card_id)
        self._validate_energy_cost(row, card_id)

        # Value validation
        self._validate_damage_values(row, card_id)
        self._validate_description_placeholders(row, card_id)

        # Action validation
        self._validate_actions(row, card_id)

        # Logic validation
        self._validate_logic_consistency(row, card_id)

        return self.errors

    def _check_required(self, row: pd.Series, field: str, card_id: str):
        if field not in row or pd.isna(row[field]) or str(row[field]).strip() == '':
            self.errors.append(ValidationError(
                card_id=card_id,
                field=field,
                message=f"Required field '{field}' is empty",
                severity="ERROR"
            ))

    def _validate_card_type(self, row: pd.Series, card_id: str):
        if 'card_type' not in row or pd.isna(row['card_type']):
            return

        card_type = row['card_type']
        valid_types = [t.value for t in CardType]

        try:
            card_type_val = int(card_type)
            if card_type_val not in valid_types:
                self.errors.append(ValidationError(
                    card_id=card_id,
                    field='card_type',
                    message=f"Invalid card_type '{card_type}'. Must be one of: {valid_types} (ATTACK=0, SKILL=1, POWER=2, STATUS=3, CURSE=4)",
                    severity="ERROR"
                ))
        except (ValueError, TypeError):
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_type',
                message=f"card_type must be a number, got: {card_type}",
                severity="ERROR"
            ))

    def _validate_rarity(self, row: pd.Series, card_id: str):
        if 'card_rarity' not in row or pd.isna(row['card_rarity']):
            return

        rarity = row['card_rarity']
        valid_rarities = [r.value for r in CardRarity]

        try:
            rarity_val = int(rarity)
            if rarity_val not in valid_rarities:
                self.errors.append(ValidationError(
                    card_id=card_id,
                    field='card_rarity',
                    message=f"Invalid card_rarity '{rarity}'. Must be one of: {valid_rarities} (BASIC=0, COMMON=1, UNCOMMON=2, RARE=3, GENERATED=4)",
                    severity="ERROR"
                ))
        except (ValueError, TypeError):
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_rarity',
                message=f"card_rarity must be a number, got: {rarity}",
                severity="ERROR"
            ))

    def _validate_color(self, row: pd.Series, card_id: str):
        if 'card_color_id' not in row or pd.isna(row['card_color_id']):
            return

        color = row['card_color_id']
        if color not in VALID_COLORS:
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_color_id',
                message=f"Unknown color '{color}'. Valid colors: {VALID_COLORS}",
                severity="WARNING"
            ))

    def _validate_energy_cost(self, row: pd.Series, card_id: str):
        if 'card_energy_cost' not in row or pd.isna(row['card_energy_cost']):
            return

        try:
            cost = int(row['card_energy_cost'])
            if cost < 0 or cost > 10:
                self.errors.append(ValidationError(
                    card_id=card_id,
                    field='card_energy_cost',
                    message=f"Energy cost {cost} seems unusual (expected 0-10)",
                    severity="WARNING"
                ))
        except (ValueError, TypeError):
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_energy_cost',
                message="Energy cost must be a number",
                severity="ERROR"
            ))

    def _validate_damage_values(self, row: pd.Series, card_id: str):
        """Validate that damage values are reasonable"""
        card_type = row.get('card_type', -1)

        # Check if it's an attack card
        try:
            is_attack = int(card_type) == CardType.ATTACK
        except:
            is_attack = False

        damage = row.get('damage', 0)
        if pd.notna(damage):
            try:
                damage_val = int(damage)
                if is_attack and damage_val == 0:
                    self.errors.append(ValidationError(
                        card_id=card_id,
                        field='damage',
                        message="Attack card has 0 damage",
                        severity="WARNING"
                    ))
                if damage_val > 100:
                    self.errors.append(ValidationError(
                        card_id=card_id,
                        field='damage',
                        message=f"Damage value {damage_val} seems very high",
                        severity="WARNING"
                    ))
            except (ValueError, TypeError):
                pass

    def _validate_description_placeholders(self, row: pd.Series, card_id: str):
        """Check that description placeholders match card_values"""
        if 'card_description' not in row or pd.isna(row['card_description']):
            return

        description = str(row['card_description'])

        # Extract placeholders like [damage], [block], etc.
        import re
        placeholders = re.findall(r'\[([a-zA-Z_]+)\]', description)

        # Known values that should be defined
        value_fields = ['damage', 'block', 'number_of_attacks', 'draw_count',
                       'status_charge_amount', 'status_secondary_charge_amount',
                       'money_amount', 'heal_amount']

        for placeholder in placeholders:
            # Skip special placeholders
            if placeholder in ['X', 'x']:
                continue

            # Check if this value is defined in the row
            if placeholder not in row or pd.isna(row[placeholder]):
                self.errors.append(ValidationError(
                    card_id=card_id,
                    field='card_description',
                    message=f"Placeholder '[{placeholder}]' in description but no value defined",
                    severity="WARNING"
                ))

    def _validate_actions(self, row: pd.Series, card_id: str):
        """Validate action scripts"""
        if 'card_play_actions' not in row or pd.isna(row['card_play_actions']):
            return

        actions_str = str(row['card_play_actions'])

        # Simple check for action references
        for action in VALID_ACTIONS:
            if action in actions_str:
                return

        # If actions column is not empty but no known action found
        if actions_str.strip() and actions_str.strip().lower() not in ['nan', 'none', '']:
            # Check if it looks like a JSON or action reference
            if not any(x in actions_str for x in ['{', '}', ':', 'Scripts.']):
                self.errors.append(ValidationError(
                    card_id=card_id,
                    field='card_play_actions',
                    message=f"Action format may be invalid: {actions_str[:50]}...",
                    severity="WARNING"
                ))

    def _validate_logic_consistency(self, row: pd.Series, card_id: str):
        """Validate logical consistency of card configuration"""
        # Check exhaust flags
        card_type = row.get('card_type', -1)
        is_power = False
        try:
            is_power = int(card_type) == CardType.POWER
        except:
            pass

        # Power cards shouldn't have exhaust flag (they're non-reusable by default)
        exhausts = row.get('card_exhausts', False)
        if is_power and exhausts:
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_exhausts',
                message="Power cards are already non-reusable, card_exhausts is redundant",
                severity="WARNING"
            ))

        # Variable cost cards
        is_variable = row.get('card_energy_cost_is_variable', False)
        energy_cost = row.get('card_energy_cost', 0)
        if is_variable and energy_cost != 0:
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_energy_cost',
                message="Variable cost cards should have energy_cost = 0",
                severity="WARNING"
            ))

        # Check target requirements
        requires_target = row.get('card_requires_target', True)
        damage = row.get('damage', 0)
        if not requires_target and pd.notna(damage) and int(damage) > 0:
            self.errors.append(ValidationError(
                card_id=card_id,
                field='card_requires_target',
                message="Card has damage but requires_target is False",
                severity="ERROR"
            ))


class CardConverter:
    """Converts Excel data to JSON format"""

    def __init__(self):
        self.validator = CardValidator()
        self.all_errors: List[ValidationError] = []

    def convert(self) -> Tuple[bool, List[ValidationError]]:
        """Main conversion function"""
        print("=" * 60)
        print("Excel to JSON Card Converter")
        print("=" * 60)

        # Check Excel file exists
        if not EXCEL_FILE.exists():
            print(f"ERROR: Excel file not found: {EXCEL_FILE}")
            print("Please create the Excel file first.")
            return False, []

        # Create output directory
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Read Excel
        try:
            print(f"\nReading Excel: {EXCEL_FILE}")
            df = pd.read_excel(EXCEL_FILE, sheet_name="Cards")
            print(f"Found {len(df)} cards")
        except Exception as e:
            print(f"ERROR reading Excel: {e}")
            return False, []

        # Validate and convert each card
        success_count = 0
        self.all_errors = []
        skipped_rows = 0

        for idx, row in df.iterrows():
            card_id = str(row.get('object_id', f'ROW_{idx}')).strip()

            # 跳过说明行（以#开头、空值、或包含"说明"/"注释"关键字）
            if not card_id or card_id.startswith('#') or card_id in ['说明', '注释', 'COMMENT', '备注', 'NOTE']:
                skipped_rows += 1
                continue

            print(f"\n[{idx+1}/{len(df)}] Processing: {card_id}")

            # Validate
            errors = self.validator.validate_card(row)
            self.all_errors.extend(errors)

            if errors:
                for err in errors:
                    prefix = "[WARN]" if err.severity == "WARNING" else "[ERROR]"
                    print(f"  {prefix}: [{err.field}] {err.message}")

            # Skip cards with critical errors
            critical_errors = [e for e in errors if e.severity == "ERROR"]
            if critical_errors:
                print(f"  [SKIP] Skipped due to {len(critical_errors)} error(s)")
                continue

            # Convert to JSON
            try:
                card_json = self._build_card_json(row)
                output_path = OUTPUT_DIR / f"{card_id}.json"

                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(card_json, f, indent=4, ensure_ascii=False)

                print(f"  [OK] Generated: {output_path}")
                success_count += 1

            except Exception as e:
                print(f"  [FAIL] Conversion failed: {e}")
                self.all_errors.append(ValidationError(
                    card_id=card_id,
                    field="CONVERSION",
                    message=str(e),
                    severity="ERROR"
                ))

        # Summary
        actual_cards = len(df) - skipped_rows
        print("\n" + "=" * 60)
        print(f"转换完成!")
        print(f"  成功: {success_count}/{actual_cards} 卡牌")
        if skipped_rows > 0:
            print(f"  跳过: {skipped_rows} 说明行")
        print(f"  错误: {len([e for e in self.all_errors if e.severity == 'ERROR'])}")
        print(f"  警告: {len([e for e in self.all_errors if e.severity == 'WARNING'])}")
        print("=" * 60)

        return success_count > 0, self.all_errors

    def _build_card_json(self, row: pd.Series) -> Dict[str, Any]:
        """Build JSON structure from Excel row"""
        card_id = str(row['object_id'])

        # Build card_values from known fields
        card_values = {}
        value_fields = [
            'damage', 'number_of_attacks', 'block', 'draw_count',
            'status_charge_amount', 'status_secondary_charge_amount',
            'money_amount', 'heal_amount', 'damage_random',
            'random_consumable', 'fill_all_slots', 'ignored_interceptor_ids',
            'status_effect_object_id', 'status_force_apply_new_effect',
            'multiplier_offset'
        ]

        for field in value_fields:
            if field in row and pd.notna(row[field]):
                val = row[field]
                # Handle boolean strings
                if isinstance(val, str):
                    if val.lower() in ('true', 'yes'):
                        card_values[field] = True
                    elif val.lower() in ('false', 'no'):
                        card_values[field] = False
                    else:
                        try:
                            card_values[field] = int(val)
                        except ValueError:
                            card_values[field] = val
                elif isinstance(val, (int, float)):
                    # Convert numbers to int to avoid float in JSON
                    try:
                        card_values[field] = int(val)
                    except (ValueError, TypeError):
                        card_values[field] = val
                else:
                    card_values[field] = val

        # Build card_play_actions
        actions = self._parse_actions(row)

        # Build upgrade improvements
        upgrade_improvements = {}
        if 'upgrade_damage' in row and pd.notna(row['upgrade_damage']):
            upgrade_improvements['damage'] = int(row['upgrade_damage'])
        if 'upgrade_block' in row and pd.notna(row['upgrade_block']):
            upgrade_improvements['block'] = int(row['upgrade_block'])
        if 'upgrade_number_of_attacks' in row and pd.notna(row['upgrade_number_of_attacks']):
            upgrade_improvements['number_of_attacks'] = int(row['upgrade_number_of_attacks'])
        if 'upgrade_draw_count' in row and pd.notna(row['upgrade_draw_count']):
            upgrade_improvements['draw_count'] = int(row['upgrade_draw_count'])
        if 'upgrade_status_charge_amount' in row and pd.notna(row['upgrade_status_charge_amount']):
            upgrade_improvements['status_charge_amount'] = int(row['upgrade_status_charge_amount'])
        if 'upgrade_status_secondary_charge_amount' in row and pd.notna(row['upgrade_status_secondary_charge_amount']):
            upgrade_improvements['status_secondary_charge_amount'] = int(row['upgrade_status_secondary_charge_amount'])
        if 'upgrade_damage_random' in row and pd.notna(row['upgrade_damage_random']):
            upgrade_improvements['damage_random'] = int(row['upgrade_damage_random'])
        if 'upgrade_multiplier_offset' in row and pd.notna(row['upgrade_multiplier_offset']):
            upgrade_improvements['multiplier_offset'] = int(row['upgrade_multiplier_offset'])

        # Build first upgrade property changes
        first_upgrade_changes = {}
        if 'first_upgrade_energy_cost' in row and pd.notna(row['first_upgrade_energy_cost']):
            first_upgrade_changes['card_energy_cost'] = int(row['first_upgrade_energy_cost'])

        # Build keywords list
        keywords = []
        if 'card_keyword_object_ids' in row and pd.notna(row['card_keyword_object_ids']):
            kw_str = str(row['card_keyword_object_ids'])
            keywords = [k.strip() for k in kw_str.split(',') if k.strip()]

        # Build the final JSON
        json_data = {
            "patch_data": {},
            "properties": {
                "object_id": card_id,
                "object_uid": "",
                "card_name": str(row.get('card_name', '')),
                "card_description": str(row.get('card_description', '')),
                "card_type": int(row.get('card_type', 0)),
                "card_rarity": int(row.get('card_rarity', 1)),
                "card_color_id": str(row.get('card_color_id', 'color_white')),
                "card_energy_cost": int(row.get('card_energy_cost', 1)),
                "card_energy_cost_is_variable": self._get_bool_or_default(row, 'card_energy_cost_is_variable', False),
                "card_energy_cost_variable_upper_bound": self._get_int_or_default(row, 'card_energy_cost_variable_upper_bound', -1),
                "card_requires_target": self._get_bool_or_default(row, 'card_requires_target', True),
                "card_exhausts": self._get_bool_or_default(row, 'card_exhausts', False),
                "card_is_ethereal": self._get_bool_or_default(row, 'card_is_ethereal', False),
                "card_is_retained": self._get_bool_or_default(row, 'card_is_retained', False),
                "card_appears_in_card_packs": self._get_bool_or_default(row, 'card_appears_in_card_packs', True),
                "card_texture_path": self._get_str_or_default(row, 'card_texture_path'),
                "card_keyword_object_ids": keywords,
                "card_values": card_values,
                "card_play_actions": actions,
                "card_upgrade_value_improvements": upgrade_improvements,
                "card_first_upgrade_property_changes": first_upgrade_changes,
                "card_upgrade_amount": 0,
                "card_upgrade_amount_max": int(row.get('card_upgrade_amount_max', 1)),
                "card_first_shuffle_priority": self._get_int_or_default(row, 'card_first_shuffle_priority', 0),
            }
        }

        return json_data

    def _get_int_or_default(self, row: pd.Series, field: str, default: int) -> int:
        if field in row and pd.notna(row[field]):
            try:
                return int(row[field])
            except (ValueError, TypeError):
                return default
        return default

    def _get_str_or_default(self, row: pd.Series, field: str, default: str = '') -> str:
        """获取字符串值，处理NaN情况"""
        if field in row and pd.notna(row[field]):
            val = str(row[field])
            # 排除pandas的nan字符串表示
            if val.lower() == 'nan':
                return default
            return val
        return default

    def _get_bool_or_default(self, row: pd.Series, field: str, default: bool = False) -> bool:
        """获取布尔值，处理NaN情况"""
        if field not in row or pd.isna(row[field]):
            return default
        val = row[field]
        if isinstance(val, bool):
            return val
        if isinstance(val, str):
            return val.lower() in ('true', 'yes', '1')
        # 对于数字，非0为True
        return bool(val)

    # Action path mappings
    ACTION_PATHS = {
        # meta_actions
        'ActionAttackGenerator': 'meta_actions',
        'ActionCardPlay': 'meta_actions',
        'ActionCardPlayEnd': 'meta_actions',
        'ActionDrawGenerator': 'meta_actions',
        'ActionEmitCustomSignal': 'meta_actions',
        'ActionValidator': 'meta_actions',
        'ActionVariableCardsetModifier': 'meta_actions',
        'ActionVariableCombatStatsModifier': 'meta_actions',
        'ActionVariableCostModifier': 'meta_actions',
        # artifact_actions
        'ActionIncreaseArtifactCharge': 'artifact_actions',
        # cardset_actions
        'ActionAddCardsToDeck': 'cardset_actions',
        'ActionAddCardsToDraw': 'cardset_actions',
        'ActionAddCardsToHand': 'cardset_actions',
        'ActionAttachCardsOntoEnemy': 'cardset_actions',
        'ActionBanishCards': 'cardset_actions',
        'ActionAttachCardToSlot': 'cardset_actions',
        'ActionChangeCardEnergies': 'cardset_actions',
        'ActionChangeCardProperties': 'cardset_actions',
        'ActionDiscardCards': 'cardset_actions',
        'ActionExhaustCards': 'cardset_actions',
        'ActionImproveCardValues': 'cardset_actions',
        'ActionMoveCardsToLimbo': 'cardset_actions',
        'ActionPlayCards': 'cardset_actions',
        'ActionRandomizeCardEnergies': 'cardset_actions',
        'ActionRemoveCardsFromDeck': 'cardset_actions',
        'ActionRetainCards': 'cardset_actions',
        'ActionTransformCards': 'cardset_actions',
        'ActionUpgradeCards': 'cardset_actions',
        # custom_ui_actions
        'ActionCustomUI': 'custom_ui_actions',
        # debug_actions
        'ActionDebugLog': 'debug_actions',
        # enemy_actions
        'ActionCycleEnemyIntent': 'enemy_actions',
        # pick_card_actions
        'ActionBasePickCards': 'pick_card_actions',
        'ActionCreateCards': 'pick_card_actions',
        'ActionPickCards': 'pick_card_actions',
        'ActionPickDuplicateCards': 'pick_card_actions',
        'ActionPickUpgradeCards': 'pick_card_actions',
        # player_actions
        'ActionAddArtifact': 'player_actions',
        'ActionAddConsumable': 'player_actions',
        'ActionAddMoney': 'player_actions',
        'ActionSwapBossArtifact': 'player_actions',
        'ActionUpdateCardDrafts': 'player_actions',
        'ActionUpdateRestActions': 'player_actions',
        'ActionUseConsumable': 'player_actions',
        # status_actions
        'ActionApplyStatus': 'status_actions',
        'ActionCorrosion': 'status_actions',
        'ActionDecayStatus': 'status_actions',
        # world_interaction_actions
        'ActionOpenChest': 'world_interaction_actions',
        'ActionStartCombat': 'world_interaction_actions',
        'ActionVisitLocation': 'world_interaction_actions',
        # generated_actions
        'ActionAttack': 'generated_actions',
        'ActionDraw': 'generated_actions',
        # rewards
        'ActionClearRewards': 'rewards',
        'ActionGrantRewards': 'rewards',
        # shop_actions
        'ActionShopPopulateItems': 'shop_actions',
        'ActionShopPurchaseItems': 'shop_actions',
        # world_generation_actions
        'ActionGenerateAct': 'world_generation_actions',
    }

    def _get_action_path(self, action_type: str) -> str:
        """Get the correct path for an action type"""
        if not action_type.startswith("Action"):
            return f"res://scripts/actions/{action_type}.gd"

        subdir = self.ACTION_PATHS.get(action_type, '')
        if subdir:
            return f"res://scripts/actions/{subdir}/{action_type}.gd"
        else:
            return f"res://scripts/actions/{action_type}.gd"

    def _parse_actions(self, row: pd.Series) -> List[Dict]:
        """Parse action configuration from row"""
        actions = []

        # Get action type from row
        action_type = row.get('action_type', '')
        if pd.isna(action_type) or str(action_type).strip() == '':
            return actions

        action_type = str(action_type).strip()

        # Build action parameters
        params = {}

        # Common parameters
        if 'action_time_delay' in row and pd.notna(row['action_time_delay']):
            params['time_delay'] = float(row['action_time_delay'])

        if 'action_target_override' in row and pd.notna(row['action_target_override']):
            to_val = row['action_target_override']
            # Handle enum string values
            if isinstance(to_val, str):
                if to_val.startswith('BaseAction.TARGET_OVERRIDES.'):
                    # Map enum names to integer values
                    target_overrides_map = {
                        'SELECTED_TARGETS': 0,
                        'PARENT': 1,
                        'PLAYER': 2,
                        'ALL_COMBATANTS': 3,
                        'ALL_ENEMIES': 4,
                        'LEFTMOST_ENEMY': 5,
                        'ENEMY_ID': 6,
                        'RANDOM_ENEMY': 7,
                    }
                    enum_name = to_val.replace('BaseAction.TARGET_OVERRIDES.', '')
                    if enum_name in target_overrides_map:
                        params['target_override'] = target_overrides_map[enum_name]
                    else:
                        params['target_override'] = to_val
                else:
                    # Try to convert to int if it's a number string
                    try:
                        params['target_override'] = int(to_val)
                    except ValueError:
                        params['target_override'] = to_val
            elif isinstance(to_val, (int, float)):
                params['target_override'] = int(to_val)
            else:
                params['target_override'] = to_val

        # Action-specific parameters
        # Handle on-lethal actions for any action type
        if 'action_on_lethal' in row and pd.notna(row['action_on_lethal']):
            lethal_action = str(row['action_on_lethal']).strip()
            if lethal_action and lethal_action != 'nan':
                lethal_action_path = self._get_action_path(lethal_action)
                params['actions_on_lethal'] = [{lethal_action_path: {}}]
        else:
            params['actions_on_lethal'] = []

        if action_type == "ActionApplyStatus":
            pass  # Uses card_values for status info

        elif action_type == "ActionValidator":
            # Complex validator structure
            if 'validator_type' in row and pd.notna(row['validator_type']):
                validator_type = str(row['validator_type'])
                params['validator_data'] = [{f"res://scripts/validators/{validator_type}.gd": {}}]

                # Passed/failed actions
                params['passed_action_data'] = []
                params['failed_action_data'] = []

        # Build final action with correct path
        action_path = self._get_action_path(action_type)
        if action_type.startswith("Action"):
            actions.append({action_path: params})

        return actions


def create_sample_excel():
    """Create a sample Excel file with proper structure"""
    print("\n创建示例Excel文件...")
    print(f"输出位置: {EXCEL_FILE}")

    # Define columns
    columns = [
        # Basic Info
        'object_id', 'card_name', 'card_description', 'card_type', 'card_rarity',
        'card_color_id', 'card_energy_cost', 'card_energy_cost_is_variable',
        'card_energy_cost_variable_upper_bound', 'card_requires_target',

        # Flags
        'card_exhausts', 'card_is_ethereal', 'card_is_retained',
        'card_appears_in_card_packs',

        # Visual
        'card_texture_path', 'card_keyword_object_ids',

        # Card Values
        'damage', 'number_of_attacks', 'block', 'draw_count',
        'status_charge_amount', 'status_secondary_charge_amount',
        'money_amount', 'heal_amount', 'damage_random',
        'status_effect_object_id', 'status_force_apply_new_effect',
        'random_consumable', 'fill_all_slots', 'ignored_interceptor_ids',
        'multiplier_offset',

        # Actions
        'action_type', 'action_time_delay', 'action_target_override',
        'action_on_lethal', 'validator_type',

        # Upgrades
        'card_upgrade_amount_max', 'first_upgrade_energy_cost',
        'upgrade_damage', 'upgrade_block', 'upgrade_number_of_attacks',
        'upgrade_draw_count', 'upgrade_status_charge_amount',
        'upgrade_status_secondary_charge_amount', 'upgrade_damage_random',
        'upgrade_multiplier_offset',

        # Shuffle
        'card_first_shuffle_priority',
    ]

    # Sample data - 第1行为说明行（会被跳过）
    sample_data = [
        {
            'object_id': '# 说明行（此行会被自动跳过）',
            'card_name': '在下方添加你的卡牌数据',
            'card_description': 'card_type: 0=攻击,1=技能,2=能力,3=状态,4=诅咒 | card_rarity: 0=基础,1=普通,2=稀有,3=罕见,4=生成',
        },
        {
            'object_id': 'card_attack_basic',
            'card_name': 'Basic Attack',
            'card_description': 'Attack for [damage] damage [number_of_attacks] times',
            'card_type': 0,  # ATTACK
            'card_rarity': 0,  # BASIC
            'card_color_id': 'color_white',
            'card_energy_cost': 1,
            'card_requires_target': True,
            'damage': 25,
            'number_of_attacks': 1,
            'action_type': 'ActionAttackGenerator',
            'action_time_delay': 0.0,
            'upgrade_damage': 1,
            'upgrade_number_of_attacks': 1,
            'card_upgrade_amount_max': 1,
        },
        {
            'object_id': 'card_block_basic',
            'card_name': 'Basic Block',
            'card_description': 'Add [block] block',
            'card_type': 1,  # SKILL
            'card_rarity': 0,  # BASIC
            'card_color_id': 'color_white',
            'card_energy_cost': 1,
            'card_requires_target': False,
            'card_keyword_object_ids': 'keyword_block',
            'block': 5,
            'action_type': 'ActionBlock',
            'action_time_delay': 0.5,
            'action_target_override': 'BaseAction.TARGET_OVERRIDES.PARENT',
            'upgrade_block': 3,
            'card_upgrade_amount_max': 1,
        },
        {
            'object_id': 'card_draw',
            'card_name': 'Draw',
            'card_description': 'Draw [draw_count] cards',
            'card_type': 1,  # SKILL
            'card_rarity': 1,  # COMMON
            'card_color_id': 'color_blue',
            'card_texture_path': 'external/sprites/cards/blue/card_blue.png',
            'card_energy_cost': 1,
            'card_requires_target': False,
            'draw_count': 3,
            'action_type': 'ActionDrawGenerator',
            'upgrade_draw_count': 1,
            'card_upgrade_amount_max': 1,
        },
    ]

    # Create DataFrame
    df = pd.DataFrame(sample_data, columns=columns)

    # Write Excel with formatting
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cards', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Cards']

        # Add data validation and formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"[OK] 示例Excel已创建: {EXCEL_FILE}")
    print(f"     包含 {len(sample_data)-1} 张示例卡牌 (第1行为说明行)")
    print("\n下一步:")
    print("  1. 使用Excel编辑卡牌配置")
    print("  2. 运行: python external/tools/convert.py")
    print("  或双击: external/tools/convert.bat")


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Excel to JSON Card Converter')
    parser.add_argument('--create-sample', action='store_true',
                       help='Create a sample Excel file')
    parser.add_argument('--validate-only', action='store_true',
                       help='Only validate, do not generate JSON')

    args = parser.parse_args()

    if args.create_sample:
        create_sample_excel()
        return

    # Run conversion
    converter = CardConverter()
    success, errors = converter.convert()

    if not success and not errors:
        # No Excel file, offer to create sample
        print("\nNo Excel file found. Create a sample?")
        print(f"  python {sys.argv[0]} --create-sample")
        sys.exit(1)

    # Exit with error code if there were errors
    critical_errors = [e for e in errors if e.severity == "ERROR"]
    sys.exit(0 if len(critical_errors) == 0 else 1)


if __name__ == "__main__":
    main()
